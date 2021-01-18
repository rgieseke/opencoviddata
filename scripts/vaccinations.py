from datetime import datetime
from pathlib import Path

import pandas as pd
import pytz
from openpyxl import load_workbook

# https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/Impfquoten-Tab.html

root = Path(__file__).parents[1]

excel_file = root / "raw/Impfquotenmonitoring.xlsx"


meta = [
    "Digitales Impfquotenmonitoring zur COVID-19-Impfung",
    "CSV-Version der Excel-Datei des RKI",
    "Quelle: https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/Impfquoten-Tab.htm",
    f"Abfragezeitpunkt: { datetime.now(pytz.timezone('Europe/Berlin')).strftime('%Y-%m-%d %H:%M:%S') }",
    "",
]

wb = load_workbook(excel_file)
meta_sheet = wb[wb.sheetnames[0]]


def iter_rows(ws):
    for row in ws.iter_rows():
        yield [
            cell.value
            for cell in row
            if (not pd.isnull(cell.value)) and (not cell.value == "")
        ]


notes = list(iter_rows(meta_sheet))
for row in notes:
    if len(row) > 0:
        meta += [" ".join(row)]

df = pd.read_excel(
    excel_file, sheet_name=1, usecols="A:F,H:I", skiprows=3, nrows=17, header=None
)
df.columns = [
    "RS",
    "Bundesland",
    "Erstimpfungen kumulativ Gesamt",
    "BioNTech",
    "Moderna",
    "Differenz zum Vortag",
    "Zweitimpfungen kumulativ Gesamt",
    "Differenz zum Vortag",
]
df["RS"] = [f"{i:02d}" if not pd.isnull(i) else "" for i in df["RS"].astype("Int64")]
df = df.set_index(["RS", "Bundesland"])

footnotes = wb[wb.sheetnames[1]]
for row in list(iter_rows(footnotes))[21:]:
    if len(row) > 0:
        meta += [" ".join(row)]

# Remove duplicates in metadata
meta = list(dict.fromkeys(meta))

with open(root / "data/vaccinations/vaccinations-total.csv", "w") as f:
    for line in meta:
        f.write(f"# {line}\n")
    f.write(df.to_csv())

# Indicators
df_indication = pd.read_excel(excel_file, sheet_name=2, skiprows=1, nrows=17)
df_indication.columns = [
    "RS",
    "Bundesland",
    "Erstimpfung - Indikation nach Alter",
    "Erstimpfung - Berufliche Indikation",
    "Erstimpfung - Medizinische Indikation",
    "Erstimpfung - PflegeheimbewohnerIn",
    "Zweitimpfung - Indikation nach Alter",
    "Zweitimpfung - Berufliche Indikation",
    "Zweitimpfung - Medizinische Indikation",
    "Zweitimpfung - PflegeheimbewohnerIn",
]

df_indication["RS"] = [
    f"{i:02d}" if not pd.isnull(i) else "" for i in df_indication["RS"].astype("Int64")
]
df_indication = df_indication.set_index(["RS", "Bundesland"])
df_indication = df_indication.astype("Int64")

footnotes = wb[wb.sheetnames[2]]
for row in list(iter_rows(footnotes))[20:]:
    if len(row) > 0:
        meta += [" ".join(row)]

# Remove duplicates in metadata
meta = list(dict.fromkeys(meta))

with open(root / "data/vaccinations/vaccinations-indication.csv", "w") as f:
    for line in meta:
        f.write(f"# {line}\n")
    f.write(df_indication.to_csv())
